import streamlit as st
import pandas as pd
import json
import io
from datetime import date, datetime
import gspread
from google.oauth2.service_account import Credentials

# ──────────────────────────────────────────────
# CONFIGURACIÓN
# ──────────────────────────────────────────────
MESAS = [1, 2, 3, 4, 5]
SCOPES = ["https://www.googleapis.com/auth/spreadsheets",
          "https://www.googleapis.com/auth/drive"]

# Nombres de las hojas dentro del Google Sheets
HOJA_INVENTARIO = "inventario"
HOJA_REGISTROS  = "registros"

# ──────────────────────────────────────────────
# CONEXIÓN A GOOGLE SHEETS
# ──────────────────────────────────────────────
@st.cache_resource
def get_client():
    creds_dict = st.secrets["gcp_service_account"]
    creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
    return gspread.authorize(creds)

def get_spreadsheet():
    gc = get_client()
    return gc.open_by_key(st.secrets["spreadsheet_id"])

def get_hoja(nombre: str):
    sh = get_spreadsheet()
    try:
        return sh.worksheet(nombre)
    except gspread.WorksheetNotFound:
        return sh.add_worksheet(title=nombre, rows=1000, cols=20)

# ──────────────────────────────────────────────
# INVENTARIO
# ──────────────────────────────────────────────
@st.cache_data(ttl=60)
def cargar_inventario() -> pd.DataFrame:
    try:
        ws = get_hoja(HOJA_INVENTARIO)
        data = ws.get_all_records()
        if not data:
            return pd.DataFrame(columns=["CÓDIGO", "INSUMO", "UM"])
        df = pd.DataFrame(data)
        df.columns = [c.strip().upper() for c in df.columns]
        return df
    except Exception as e:
        st.error(f"Error cargando inventario: {e}")
        return pd.DataFrame(columns=["CÓDIGO", "INSUMO", "UM"])

def guardar_inventario(df: pd.DataFrame):
    try:
        ws = get_hoja(HOJA_INVENTARIO)
        ws.clear()
        df.columns = [c.upper() for c in df.columns]
        ws.update([df.columns.tolist()] + df.fillna("").values.tolist())
        cargar_inventario.clear()
    except Exception as e:
        st.error(f"Error guardando inventario: {e}")
        st.stop()

# ──────────────────────────────────────────────
# REGISTROS DE CONTEO
# ──────────────────────────────────────────────
@st.cache_data(ttl=30)
def cargar_registros() -> dict:
    try:
        ws  = get_hoja(HOJA_REGISTROS)
        data = ws.get_all_records()
        registros = {}
        def safe_float(val):
            if isinstance(val, str):
                val = val.replace(",", ".")
            try:
                return float(val or 0)
            except:
                return 0.0

        for row in data:
            clave = f"{row.get('fecha','')}__{row.get('codigo','')}"
            mesas = {}
            for m in MESAS:
                mesas[str(m)] = safe_float(row.get(f"mesa{m}", 0))
            registros[clave] = {
                "fecha":   row.get("fecha", ""),
                "codigo":  row.get("codigo", ""),
                "insumo":  row.get("insumo", ""),
                "um":      row.get("um", ""),
                "mesas":   mesas,
                "total":   safe_float(row.get("total", 0)),
                "updated": row.get("updated", ""),
            }
        return registros
    except Exception as e:
        st.error(f"Error cargando registros: {e}")
        return {}

def guardar_registros(data: dict):
    try:
        ws = get_hoja(HOJA_REGISTROS)
        headers = ["fecha","codigo","insumo","um",
                   "mesa1","mesa2","mesa3","mesa4","mesa5","total","updated"]
        rows = [headers]
        for v in data.values():
            m = v.get("mesas", {})
            rows.append([
                v.get("fecha",""), v.get("codigo",""), v.get("insumo",""), v.get("um",""),
                float(m.get("1",0)), float(m.get("2",0)), float(m.get("3",0)),
                float(m.get("4",0)), float(m.get("5",0)),
                float(v.get("total",0)), v.get("updated","")
            ])
        ws.clear()
        ws.update(rows)
        cargar_registros.clear()
    except Exception as e:
        st.error(f"Error guardando registros: {e}")

# ──────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────
def registros_a_df(data: dict) -> pd.DataFrame:
    if not data:
        return pd.DataFrame()
    rows = []
    for v in data.values():
        row = {"Fecha": v.get("fecha",""), "Código": v.get("codigo",""),
               "Insumo": v.get("insumo",""), "UM": v.get("um","")}
        for m in MESAS:
            row[f"Mesa {m}"] = v.get("mesas",{}).get(str(m), 0)
        row["Total"] = v.get("total", 0)
        rows.append(row)
    df = pd.DataFrame(rows)
    df["Fecha"] = pd.to_datetime(df["Fecha"])
    return df

def excel_bytes(df: pd.DataFrame) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook(); ws = wb.active
    color_s = "2C6FB5"
    thin  = Side(style="thin", color="CCCCCC")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)
    cols  = df.columns.tolist()
    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=color_s)
        c.alignment = Alignment(horizontal="center")
        c.border = borde
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Arial", size=10)
            c.border = borde
            c.alignment = Alignment(horizontal="left" if ci==3 else "center")
        if ri % 2 == 0:
            for ci in range(1, len(cols)+1):
                ws.cell(row=ri, column=ci).fill = PatternFill("solid", fgColor="EEF2F7")
    for ci in range(1, len(cols)+1):
        ws.column_dimensions[get_column_letter(ci)].width = 18
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()

# ──────────────────────────────────────────────
# ESTILOS
# ──────────────────────────────────────────────
st.markdown("""
<style>
    .stApp { background-color: #f5f7fa; }
    .block-container { padding-top: 1.5rem; }
    h1 { color: #1a3a5c; }
    h2, h3 { color: #2c5282; }
    .metric-box {
        background: white; border-radius: 10px; padding: 16px 20px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.08); text-align: center;
    }
    .total-box {
        background: linear-gradient(135deg,#1a3a5c,#2c6fb5); color: white;
        border-radius: 12px; padding: 20px; text-align: center;
        font-size: 2rem; font-weight: 700;
        box-shadow: 0 4px 12px rgba(44,111,181,0.3);
    }
    div[data-testid="stTabs"] button { font-weight: 600; }
</style>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────
# CABECERA
# ──────────────────────────────────────────────
st.title("🧪 Sistema de Inventario — Dosimetría")
st.markdown("---")

# ──────────────────────────────────────────────
# PESTAÑAS PRINCIPALES
# ──────────────────────────────────────────────
tab1, tab2, tab3, tab4 = st.tabs([
    "📋 Registro de Conteo",
    "🔍 Consulta por Fecha",
    "📊 Reportes",
    "⚙️ Gestión de Insumos",
])

# ══════════════════════════════════════════════
# TAB 1 — REGISTRO DE CONTEO
# ══════════════════════════════════════════════
with tab1:
    df_inv    = cargar_inventario()
    registros = cargar_registros()
    st.subheader("📋 Registro de Conteo por Mesa")

    col_fecha, col_buscar = st.columns([2, 4])
    with col_fecha:
        fecha_sel = st.date_input("📅 Fecha de conteo", value=date.today(), key="fecha_reg")
        fecha_str = fecha_sel.strftime("%Y-%m-%d")
    with col_buscar:
        insumo_sel = st.selectbox(
            "🔍 Buscar y seleccionar insumo",
            ["-- Seleccione un insumo --"] + df_inv["INSUMO"].tolist(),
            key="insumo_sel", help="Escribe directamente aquí para filtrar"
        )

    if insumo_sel != "-- Seleccione un insumo --":
        fila   = df_inv[df_inv["INSUMO"] == insumo_sel].iloc[0]
        codigo = fila["CÓDIGO"]
        um     = fila.get("UM", "")

        c1, c2, c3 = st.columns(3)
        c1.markdown(f"<div class='metric-box'>🔑 <b>Código</b><br>{codigo}</div>", unsafe_allow_html=True)
        c2.markdown(f"<div class='metric-box'>📦 <b>Insumo</b><br>{insumo_sel}</div>", unsafe_allow_html=True)
        c3.markdown(f"<div class='metric-box'>⚖️ <b>UM</b><br>{um}</div>", unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)

        clave         = f"{fecha_str}__{codigo}"
        existente     = registros.get(clave, {})
        mesas_previas = existente.get("mesas", {str(m): 0 for m in MESAS})

        st.markdown("### 🏭 Conteo por Mesa")
        cols_mesas   = st.columns(5)
        valores_mesa = {}
        for i, m in enumerate(MESAS):
            with cols_mesas[i]:
                v = st.number_input(f"Mesa {m}", min_value=0.0,
                                    value=float(mesas_previas.get(str(m), 0)),
                                    step=0.5, key=f"mesa_{m}_{clave}")
                valores_mesa[str(m)] = v

        total = sum(valores_mesa.values())
        st.markdown(f"<div class='total-box'>TOTAL DE CONTEO: {total:,.2f} {um}</div>", unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)

        col_guard, col_eliminar = st.columns([3, 1])
        with col_guard:
            if st.button("💾 Guardar Registro", use_container_width=True, type="primary"):
                registros[clave] = {
                    "fecha": fecha_str, "codigo": codigo, "insumo": insumo_sel,
                    "um": um, "mesas": valores_mesa, "total": total,
                    "updated": datetime.now().isoformat()
                }
                with st.spinner("Guardando..."):
                    guardar_registros(registros)
                st.success(f"✅ Registro guardado para **{insumo_sel}** el **{fecha_str}**")
                st.balloons()
        with col_eliminar:
            if clave in registros:
                if st.button("🗑️ Eliminar Registro", use_container_width=True, type="secondary"):
                    del registros[clave]
                    with st.spinner("Eliminando..."):
                        guardar_registros(registros)
                    st.warning(f"⚠️ Registro eliminado para **{insumo_sel}** el **{fecha_str}**")
                    st.rerun()

    st.markdown("---")
    registros_dia = [v for v in registros.values() if v.get("fecha") == fecha_str]
    if registros_dia:
        st.markdown(f"### 📑 Registros del día: {fecha_str} ({len(registros_dia)} insumos)")
        df_dia = pd.DataFrame(registros_dia)[["codigo","insumo","mesas","total","um"]]
        for m in MESAS:
            df_dia[f"Mesa {m}"] = df_dia["mesas"].apply(lambda x: x.get(str(m), 0))
        df_dia = df_dia.drop(columns=["mesas"]).rename(columns={
            "codigo":"Código","insumo":"Insumo","total":"Total","um":"UM"})
        st.dataframe(df_dia, use_container_width=True, hide_index=True)
    else:
        st.info(f"📭 No hay registros para el {fecha_str}")

# ══════════════════════════════════════════════
# TAB 2 — CONSULTA POR FECHA
# ══════════════════════════════════════════════
with tab2:
    st.subheader("🔍 Consulta de Inventario por Fecha")
    registros = cargar_registros()
    fechas_disponibles = sorted(set(v["fecha"] for v in registros.values()), reverse=True)

    if not fechas_disponibles:
        st.info("📭 Aún no hay registros guardados.")
    else:
        fecha_consulta     = st.date_input("📅 Seleccione la fecha", value=date.today(), key="fecha_consulta")
        fecha_consulta_str = fecha_consulta.strftime("%Y-%m-%d")
        resultados = [v for v in registros.values() if v.get("fecha") == fecha_consulta_str]

        if resultados:
            st.success(f"✅ **{len(resultados)}** registros para el **{fecha_consulta_str}**")
            df_res = pd.DataFrame(resultados)
            for m in MESAS:
                df_res[f"Mesa {m}"] = df_res["mesas"].apply(lambda x: x.get(str(m), 0))
            df_res = df_res.drop(columns=["mesas","updated"], errors="ignore").rename(columns={
                "fecha":"Fecha","codigo":"Código","insumo":"Insumo","total":"Total","um":"UM"})
            df_res = df_res[["Código","Insumo","UM"]+[f"Mesa {m}" for m in MESAS]+["Total"]].sort_values("Insumo")
            st.dataframe(df_res, use_container_width=True, hide_index=True)
            c1, c2 = st.columns(2)
            c1.metric("📦 Insumos contados", len(resultados))
            c2.metric("⚖️ Total general", f"{df_res['Total'].sum():,.2f}")
            st.download_button("📥 Descargar Excel", data=excel_bytes(df_res),
                               file_name=f"conteo_{fecha_consulta_str}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.warning(f"📭 No hay registros para **{fecha_consulta_str}**")

        with st.expander("📆 Ver todas las fechas con registros"):
            for f in fechas_disponibles:
                n = len([v for v in registros.values() if v["fecha"] == f])
                st.markdown(f"• **{f}** — {n} insumo(s)")

# ══════════════════════════════════════════════
# TAB 3 — REPORTES
# ══════════════════════════════════════════════
with tab3:
    st.subheader("📊 Reportes de Consumo")
    registros = cargar_registros()
    df_todos  = registros_a_df(registros)

    if df_todos.empty:
        st.info("📭 Aún no hay registros para generar reportes.")
    else:
        rep1, rep2, rep3 = st.tabs([
            "📅 Reporte Mensual",
            "📈 Gráficas de Consumo",
            "🗓️ Exportar por Rango de Fechas",
        ])

        with rep1:
            st.markdown("#### Reporte Mensual de Conteo")
            df_todos["Año-Mes"] = df_todos["Fecha"].dt.to_period("M").astype(str)
            meses_disponibles   = sorted(df_todos["Año-Mes"].unique(), reverse=True)
            mes_sel = st.selectbox("Selecciona el mes", meses_disponibles, key="mes_rep")
            df_mes  = df_todos[df_todos["Año-Mes"] == mes_sel].drop(columns=["Año-Mes"]).copy()
            df_mes  = df_mes.sort_values(["Fecha","Insumo"])
            df_mes["Fecha"] = df_mes["Fecha"].dt.strftime("%Y-%m-%d")
            st.success(f"📋 **{len(df_mes)}** registros en **{mes_sel}**")
            df_resumen = (df_mes.groupby(["Código","Insumo","UM"])["Total"]
                          .sum().reset_index().sort_values("Total", ascending=False)
                          .rename(columns={"Total":"Total del Mes"}))
            col_det, col_res = st.columns(2)
            cols_det = ["Fecha","Código","Insumo","UM"] + [f"Mesa {m}" for m in MESAS] + ["Total"]
            with col_det:
                st.markdown("**Detalle por día**")
                st.dataframe(df_mes[cols_det], use_container_width=True, hide_index=True)
            with col_res:
                st.markdown("**Resumen acumulado por insumo**")
                st.dataframe(df_resumen, use_container_width=True, hide_index=True)
            c1, c2, c3 = st.columns(3)
            c1.metric("📦 Registros", len(df_mes))
            c2.metric("🧴 Insumos distintos", df_mes["Insumo"].nunique())
            c3.metric("⚖️ Total acumulado", f"{df_mes['Total'].sum():,.2f}")
            col_d, col_r = st.columns(2)
            with col_d:
                st.download_button("📥 Descargar detalle Excel", data=excel_bytes(df_mes[cols_det]),
                                   file_name=f"detalle_{mes_sel}.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            with col_r:
                st.download_button("📥 Descargar resumen Excel", data=excel_bytes(df_resumen),
                                   file_name=f"resumen_{mes_sel}.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        with rep2:
            st.markdown("#### Gráficas de Consumo")
            try:
                import plotly.express as px
                col_g1, col_g2 = st.columns(2)
                with col_g1:
                    st.markdown("**🏆 Top 10 Insumos por cantidad total**")
                    df_top = (df_todos.groupby("Insumo")["Total"].sum().reset_index()
                              .sort_values("Total", ascending=False).head(10))
                    fig1 = px.bar(df_top, x="Total", y="Insumo", orientation="h",
                                  color="Total", color_continuous_scale=["#90caf9","#1a3a5c"],
                                  labels={"Total":"Cantidad","Insumo":""})
                    fig1.update_layout(showlegend=False, yaxis=dict(autorange="reversed"),
                                       plot_bgcolor="white", height=380, margin=dict(l=10,r=10,t=10,b=10))
                    st.plotly_chart(fig1, use_container_width=True)
                with col_g2:
                    st.markdown("**📅 Total de conteo por fecha**")
                    df_xf = (df_todos.groupby("Fecha")["Total"].sum().reset_index().sort_values("Fecha"))
                    df_xf["Fecha_str"] = df_xf["Fecha"].dt.strftime("%d/%m/%Y")
                    fig2 = px.line(df_xf, x="Fecha_str", y="Total", markers=True,
                                   color_discrete_sequence=["#2c6fb5"],
                                   labels={"Fecha_str":"Fecha","Total":"Total contado"})
                    fig2.update_traces(line_width=2.5, marker_size=8)
                    fig2.update_layout(plot_bgcolor="white", height=380, margin=dict(l=10,r=10,t=10,b=10))
                    st.plotly_chart(fig2, use_container_width=True)
                st.markdown("---")
                st.markdown("**📈 Evolución de consumo por insumo**")
                insumo_g = st.selectbox("Selecciona el insumo", sorted(df_todos["Insumo"].unique()), key="insumo_grafica")
                df_evol  = df_todos[df_todos["Insumo"] == insumo_g].sort_values("Fecha").copy()
                df_evol["Fecha_str"] = df_evol["Fecha"].dt.strftime("%d/%m/%Y")
                fig3 = px.area(df_evol, x="Fecha_str", y="Total", markers=True,
                               color_discrete_sequence=["#2c6fb5"], labels={"Fecha_str":"Fecha","Total":"Cantidad"})
                fig3.update_layout(plot_bgcolor="white", height=320, margin=dict(l=10,r=10,t=10,b=10))
                st.plotly_chart(fig3, use_container_width=True)
                st.markdown("**🏭 Distribución acumulada por Mesa**")
                totales_mesa = {f"Mesa {m}": float(df_todos[f"Mesa {m}"].sum()) for m in MESAS}
                df_mf = pd.DataFrame({"Mesa": list(totales_mesa.keys()), "Total": list(totales_mesa.values())})
                fig4  = px.pie(df_mf, names="Mesa", values="Total",
                               color_discrete_sequence=px.colors.sequential.Blues_r, hole=0.4)
                fig4.update_layout(height=350, margin=dict(l=10,r=10,t=10,b=10))
                st.plotly_chart(fig4, use_container_width=True)
            except ImportError:
                st.error("Instala plotly: `pip install plotly`")

        with rep3:
            st.markdown("#### Exportar por Rango de Fechas")
            fecha_min = df_todos["Fecha"].min().date()
            fecha_max = df_todos["Fecha"].max().date()
            col_fi, col_ff = st.columns(2)
            with col_fi:
                f_inicio = st.date_input("📅 Fecha inicio", value=fecha_min,
                                         min_value=fecha_min, max_value=fecha_max, key="f_ini")
            with col_ff:
                f_fin = st.date_input("📅 Fecha fin", value=fecha_max,
                                       min_value=fecha_min, max_value=fecha_max, key="f_fin")
            if f_inicio > f_fin:
                st.error("⚠️ La fecha inicio no puede ser mayor a la fecha fin.")
            else:
                df_rango = df_todos[
                    (df_todos["Fecha"].dt.date >= f_inicio) &
                    (df_todos["Fecha"].dt.date <= f_fin)
                ].copy().sort_values(["Fecha","Insumo"])
                df_rango["Fecha"] = df_rango["Fecha"].dt.strftime("%Y-%m-%d")
                if df_rango.empty:
                    st.warning("📭 No hay registros en ese rango.")
                else:
                    st.success(f"✅ **{len(df_rango)}** registros entre **{f_inicio}** y **{f_fin}**")
                    agrupacion = st.radio("Ver como:", ["Detalle por día", "Resumen por insumo"], horizontal=True)
                    if agrupacion == "Detalle por día":
                        cols_r  = ["Fecha","Código","Insumo","UM"] + [f"Mesa {m}" for m in MESAS] + ["Total"]
                        df_show = df_rango[cols_r]
                        nombre_archivo = f"detalle_{f_inicio}_{f_fin}.xlsx"
                    else:
                        df_show = (df_rango.groupby(["Código","Insumo","UM"])["Total"]
                                   .sum().reset_index().sort_values("Total", ascending=False)
                                   .rename(columns={"Total":"Total Acumulado"}))
                        nombre_archivo = f"resumen_{f_inicio}_{f_fin}.xlsx"
                    st.dataframe(df_show, use_container_width=True, hide_index=True)
                    c1, c2, c3 = st.columns(3)
                    c1.metric("📦 Registros", len(df_rango))
                    c2.metric("🧴 Insumos distintos", df_rango["Insumo"].nunique())
                    c3.metric("⚖️ Total acumulado", f"{df_rango['Total'].sum():,.2f}")
                    st.download_button("📥 Descargar Excel", data=excel_bytes(df_show),
                                       file_name=nombre_archivo,
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ══════════════════════════════════════════════
# TAB 4 — GESTIÓN DE INSUMOS
# ══════════════════════════════════════════════
with tab4:
    st.subheader("⚙️ Gestión de Insumos del Catálogo")
    df_inv = cargar_inventario()
    sub1, sub2, sub3 = st.tabs(["➕ Nuevo Insumo", "✏️ Editar Insumo", "🗑️ Eliminar Insumo"])

    with sub1:
        st.markdown("#### Crear nuevo insumo")
        n_codigo = st.text_input("Código *", key="n_cod")
        n_insumo = st.text_input("Nombre del insumo *", key="n_ins")
        n_um     = st.text_input("Unidad de medida (UM)", value="KILOGRAMO", key="n_um")
        if st.button("✅ Crear Insumo", type="primary"):
            if not n_codigo or not n_insumo:
                st.error("El código y el nombre son obligatorios.")
            elif n_codigo in df_inv["CÓDIGO"].values:
                st.error(f"Ya existe un insumo con el código **{n_codigo}**.")
            elif n_insumo.upper() in df_inv["INSUMO"].str.upper().values:
                st.error(f"Ya existe el insumo **{n_insumo}**.")
            else:
                nueva  = pd.DataFrame([{"CÓDIGO": n_codigo.strip(),
                                         "INSUMO": n_insumo.strip().upper(),
                                         "UM": n_um.strip().upper()}])
                df_inv = pd.concat([df_inv, nueva], ignore_index=True).sort_values("INSUMO")
                with st.spinner("Guardando..."):
                    guardar_inventario(df_inv)
                st.success(f"✅ Insumo **{n_insumo}** creado correctamente.")
                st.rerun()

    with sub2:
        st.markdown("#### Editar insumo existente")
        sel_edit = st.selectbox("🔍 Buscar y seleccionar insumo a editar",
                                 ["-- Seleccione --"] + df_inv["INSUMO"].tolist(),
                                 key="sel_edit", help="Escribe para filtrar")
        if sel_edit != "-- Seleccione --":
            fila_edit = df_inv[df_inv["INSUMO"] == sel_edit].iloc[0]
            idx_edit  = df_inv[df_inv["INSUMO"] == sel_edit].index[0]
            e_cod = st.text_input("Código", value=fila_edit["CÓDIGO"], key="e_cod")
            e_ins = st.text_input("Nombre", value=fila_edit["INSUMO"],  key="e_ins")
            e_um  = st.text_input("UM",     value=fila_edit.get("UM",""), key="e_um")
            if st.button("💾 Guardar cambios", type="primary"):
                df_inv.at[idx_edit, "CÓDIGO"] = e_cod.strip()
                df_inv.at[idx_edit, "INSUMO"] = e_ins.strip().upper()
                df_inv.at[idx_edit, "UM"]     = e_um.strip().upper()
                with st.spinner("Guardando..."):
                    guardar_inventario(df_inv)
                st.success("✅ Insumo actualizado correctamente.")
                st.rerun()

    with sub3:
        st.markdown("#### Eliminar insumo del catálogo")
        st.warning("⚠️ Esta acción eliminará el insumo permanentemente del catálogo.")
        sel_del = st.selectbox("🔍 Buscar y seleccionar insumo a eliminar",
                                ["-- Seleccione --"] + df_inv["INSUMO"].tolist(),
                                key="sel_del", help="Escribe para filtrar")
        if sel_del != "-- Seleccione --":
            fila_del = df_inv[df_inv["INSUMO"] == sel_del].iloc[0]
            st.error(f"¿Eliminar **{sel_del}** (Código: {fila_del['CÓDIGO']})?")
            if st.checkbox("Confirmo que deseo eliminar este insumo", key="confirm_del"):
                if st.button("🗑️ Eliminar definitivamente", type="primary"):
                    df_inv = df_inv[df_inv["INSUMO"] != sel_del].reset_index(drop=True)
                    with st.spinner("Eliminando..."):
                        guardar_inventario(df_inv)
                    st.success(f"✅ Insumo **{sel_del}** eliminado.")
                    st.rerun()

    st.markdown("---")
    with st.expander(f"📚 Ver catálogo completo ({len(df_inv)} insumos)"):
        st.dataframe(df_inv.sort_values("INSUMO"), use_container_width=True, hide_index=True)
